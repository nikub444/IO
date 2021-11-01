from openpyxl import Workbook, load_workbook
import test2
import random
# setup
wb = load_workbook('Dane_S2_50_10.xlsx', data_only=False)
ws = wb['Arkusz1']
iterations = 10
tasks = 50


def HillClimbing():
    for x in range(iterations):
        globalMin = test2.Min()
        # losujemy dwie różne liczby
        while True:
            rand1 = int((tasks - 1 + 1) * random.random() + 1) + 1
            rand2 = int((tasks - 1 + 1) * random.random() + 1) + 1
            # print(rand1 != rand2)
            # print(rand1, rand2)
            if (rand1 != rand2):
                # print(rand1, rand2)
                break
        cell1 = ws.cell(rand1, 13).value
        cell2 = ws.cell(rand2, 13).value
        ws.cell(rand1, 13).value = cell2
        ws.cell(rand2, 13).value = cell1
        wb.save('Dane_S2_50_10.xlsx')
        localMin = test2.Min()
        print(globalMin, localMin, globalMin < localMin)
        if(globalMin < localMin):
            ws.cell(rand1, 13).value = cell1
            ws.cell(rand2, 13).value = cell2
        wb.save('Dane_S2_50_10.xlsx')


HillClimbing()
wb.save('Dane_S2_50_10.xlsx')
